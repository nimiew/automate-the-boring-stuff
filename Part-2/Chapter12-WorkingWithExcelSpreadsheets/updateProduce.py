#! python3
# updateProduce.py - Correct costs in produce sales spreadsheet

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# Produce types and updated prices
PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27,
    }

# Loop through rows and update prices
for i in range(2, sheet.max_row):
    produceName = sheet.cell(row=i, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=i, column=1).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
